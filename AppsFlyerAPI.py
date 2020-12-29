import datetime
import requests
import urllib3
import pandas as pd
from io import StringIO

urllib3.disable_warnings()
date_start = (datetime.date.today() - datetime.timedelta(days=10)).strftime("%Y-%m-%d")
date_stop = (datetime.date.today() - datetime.timedelta(days=1)).strftime("%Y-%m-%d")

weekNumber = (datetime.date.today() - datetime.timedelta(days=2)).strftime("%Y-%V")
print(weekNumber)
filePath = r'\\srv\WEB-ANALYTICS\File_transfer\allDataForReports111.xlsx'
apps = {'id582200000':'iOS','ru.ukt.android':'Android'}


def dataFromAppsFlyer(app_id, app_name):
    with open(r'C:\Python\ukt\pass\AppsFlyer.txt') as f:
        tokenApp = f.readline()
    url = f'https://hq.appsflyer.com/export/{app_id}/partners_report/v5'
    # url = 'https://hq.appsflyer.com/export/master_report/v4'
    param = {'api_token': tokenApp, 'from': date_start, 'to': date_stop, 'timezone': 'Europe/Moscow'}
    # param = {'api_token': tokenApp, 'from': date_start, 'to': date_stop, 'app_id': app_id, 'groupings': app_id , 'kpis': 'activity_sessions'}
    response_app = requests.get(url, params=param, verify=False)
    print(response_app)
    RAWDATA = StringIO(response_app.text)
    print(RAWDATA)
    df = pd.read_csv(RAWDATA)
    df.columns = [i.replace(' (Unique users)', '') for i in list(df.columns.values)]

    df_first = df.filter(['Media Source (pid)', 'af_add_to_cart', 'af_search', 'af_opened_from_push_notification',
                                            'af_initiated_checkout', 'af_add_to_wishlist','af_rate','Impressions','Clicks','Installs'], axis=1)

    df_first = df_first.groupby(['Media Source (pid)']).sum()
    df_first = df_first.reset_index()
    df_first['system'] = app_name
    df_first['week'] = weekNumber
    df_first['ctr'] = df_first['Installs'] / df_first['Clicks']

    df_events = df_first[['week','system','Media Source (pid)','af_add_to_cart','af_search','af_initiated_checkout',
               'af_opened_from_push_notification','af_add_to_wishlist','af_rate']]

    df_metrics = df_first[['week','system','Media Source (pid)','Impressions','Clicks','Installs','ctr']]

    return df_events, df_metrics


dataFromAppsFlyer('ru.ukt.android.','android')


def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None, truncate_sheet=False, **to_excel_kwargs):
    from openpyxl import load_workbook
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')
    writer = pd.ExcelWriter(filename, engine='openpyxl')
    try:
        writer.book = load_workbook(filename)
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            idx = writer.book.sheetnames.index(sheet_name)
            writer.book.remove(writer.book.worksheets[idx])
            writer.book.create_sheet(sheet_name, idx)
        writer.sheets = {ws.title: ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        pass
    if startrow is None:
        startrow = 0
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)
    writer.save()

def main():
    events = pd.DataFrame({})
    metrics = pd.DataFrame({})
    for i,j in apps.items():
        y, z = dataFromAppsFlyer(i,j)
        events = pd.concat([y, events])
        metrics = pd.concat([z, metrics])
    append_df_to_excel(filePath, events, sheet_name='WEEK APF events', index=False, header=False)
    append_df_to_excel(filePath, metrics, sheet_name='WEEK APF metrics', index=False, header=False)

# main()