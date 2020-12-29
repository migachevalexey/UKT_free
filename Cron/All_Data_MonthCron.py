import requests
import urllib
import datetime
import pandas as pd
from calendar import monthrange
from pandas import ExcelWriter
from apiclient.discovery import build
from oauth2client.service_account import ServiceAccountCredentials
import httplib2
import urllib3
from io import StringIO

"""
Тянем данные из GA, из Яндекс AppMetrica(в разрезе приложения) и из AppsFlyer(в разрезе приложения)
Сохраняем все это в excel с накоплением данных
На ежемесячной основе.
"""

urllib3.disable_warnings()
# print(urllib.parse.unquote("metrics=ym%3Au%3AactiveUsers%2Cym%3Au%3AactiveUsersPercentage%2Cym%3Au%3AnewUsers%2Cym%3Au%3AnewUsersPercentage%2Cym%3Au%3AnewUsersShare&dimensions=ym%3Au%3Adate&sort=-ym%3Au%3Adate&offset=1&limit=50&include_undefined=true&accuracy=1&proposedAccuracy=true", encoding='utf-8'))

SCOPES = ['https://www.googleapis.com/auth/analytics.readonly']
SERVICE_ACCOUNT_EMAIL = 'ukt-cloud@konic-progress-196909.iam.gserviceaccount.com'
DISCOVERY_URI = 'https://analyticsreporting.googleapis.com/$discovery/rest'
KEY_FILE_LOCATION = r'C:\Python\ukt\pass\uktOwox-d00000000e12.p12'
VIEW_ID = '115850000'  # id "Correct Master Data"
apps = {'com.akzia.apps.ukt':'iOS','ru.ukt.android':'Android'}
appsFlyer = {'id582200000':'iOS','ru.ukt.android':'Android'}
filePath = r'\\srv\WEB-ANALYTICS\File_transfer\allDataForReports.xlsx'


curr_date = datetime.date.today()
month = curr_date.replace(month=datetime.date.today().month - 1)  # предыдущий месяц
last_day = str(monthrange(curr_date.year, curr_date.month - 1)[1])  # последний день предыдущего месяца
start_month = month.strftime("%Y-%m-01")  # начало предыдущего месяца
end_month = month.strftime("%Y-%m-") + last_day  # конец предыдущего месяца

print(start_month, end_month)


def dataFromAppMetrika(date_start, date_stop):
    d = []
    for i,j in apps.items():
        params_sess = {'lang': 'ru', 'request_domain': 'ru',
                       'filters': "exists ym:d:device with (appID=='{}')".format(i), 'id': 516000,
                       'date1': date_start, 'date2': date_stop, 'metrics': 'ym:s:sessions', 'dimensions': 'ym:s:date',
                       'sort': '-ym:s:date', 'offset': 1, 'limit': 10, 'accuracy': 1, 'proposedAccuracy': 'true',}
        response_ses = requests.get('https://api.appmetrica.yandex.ru/stat/v1/data', params=params_sess,
                                    headers={'Authorization': 'OAuth Ag'}).json()
        params_users = {'lang': 'ru', 'request_domain': 'ru',
                        'filters': "exists ym:d:device with (appID=='{}')".format(i), 'id': 516000,
                        'date1': date_start, 'date2': date_stop, 'metrics': 'ym:u:activeUsers',
                        'dimensions': 'ym:u:date', 'sort': '-ym:u:date', 'include_undefined': 'true', 'offset': 1,
                        'limit': 10, 'accuracy': 1, 'proposedAccuracy': 'true', }
        response_user = requests.get('https://api.appmetrica.yandex.ru/stat/v1/data', params=params_users,
                                     headers={'Authorization': 'OAuth Ag'}).json()
        d.append([int(month.strftime("%Y%m")),j,  int(*response_ses['totals']), int(*response_user['totals'])])

    return d

def initialize_analyticsreporting():
    credentials = ServiceAccountCredentials.from_p12_keyfile(
        SERVICE_ACCOUNT_EMAIL, KEY_FILE_LOCATION, scopes=SCOPES)
    http = credentials.authorize(httplib2.Http())
    analytics = build('analytics', 'v4', http=http, discoveryServiceUrl=DISCOVERY_URI)

    return analytics


def dataFromGA(analytics, ga_dim):
    QUERY = {
        'reportRequests': [
            {'viewId': VIEW_ID,
             'dateRanges': [{'startDate': start_month, 'endDate': end_month}],
             "samplingLevel": 'LARGE',
             "dimensions": [{"name": "ga:yearMonth"}, {"name": ga_dim}],
             'metrics': [{"expression": "ga:sessions"}, {"expression": "ga:transactions"},
                         {"expression": "ga:transactionRevenue"}, ],

             'pageSize': 10000}
        ]}
    request = analytics.reports().batchGet(body=QUERY).execute()  # запрос API
    rowCount = request['reports'][0]['data']['rowCount']
    API_data = request['reports'][0]['data']['rows']
    if rowCount > 10000:  # если записей болше 10т, упираемся в лимит и делаем цикл с параметром nextPageToken
        for i in range(rowCount // 10000):
            pageToken = request['reports'][0]["nextPageToken"]
            QUERY['reportRequests'][0].update({"pageToken": pageToken})  # добавляем параметр pageToken в dict запроса
            request = analytics.reports().batchGet(body=QUERY).execute()  # посылаем запрос API
            API_data += request['reports'][0]['data']['rows']

    print(f'Колво записей: {rowCount}')

    API_data = [list(i['dimensions']) + list(i['metrics'][0]['values']) for i in API_data]
    for i in API_data:
        i[4] = str(int(float(i[4])))
    API_data = [[int(i) if i.isdigit() else i for i in j] for j in API_data]

    return API_data


def dataFromAppsFlyer(app_id, app_name):
    with open(r'C:\Python\ukt\pass\AppsFlyer.txt') as f:
        tokenApp = f.readline()
    url = f'https://hq.appsflyer.com/export/{app_id}/partners_report/v5'
    param = {'api_token': tokenApp, 'from': start_month, 'to': end_month, 'timezone': 'Europe/Moscow'}
    response_app = requests.get(url, params=param, verify=False)
    RAWDATA = StringIO(response_app.text)

    df = pd.read_csv(RAWDATA)
    df.columns = [i.replace(' (Unique users)', '') for i in list(df.columns.values)]
    # df_events = df.filter(['Media Source (pid)', 'af_add_to_cart', 'af_search', 'af_opened_from_push_notification'], axis=1)

    df_first = df.filter(['Media Source (pid)', 'af_rate','Impressions','Clicks','Installs'], axis=1)

    df_first = df_first.groupby(['Media Source (pid)']).sum().reset_index()
    df_first['system'] = app_name
    df_first['month'] = month.strftime("%Y-%m")
    df_first['ctr'] = df_first['Installs'] / df_first['Clicks']


    df_metrics = df_first[['month','system','Media Source (pid)','Impressions','Clicks','Installs','ctr']]

    return df_metrics


def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None, truncate_sheet=False, **to_excel_kwargs):
    from openpyxl import load_workbook
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')
    writer = pd.ExcelWriter(filename, engine='openpyxl')
    try:
        writer.book = load_workbook(filename)
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            idx = writer.book.sheetnames.index(sheet_name)
            writer.book.remove(writer.book.worksheets[idx])
            writer.book.create_sheet(sheet_name, idx)
        writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        pass
    if startrow is None:
        startrow = 0
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)
    writer.save()
    writer.close()


def main():
    analytics = initialize_analyticsreporting()
    data_app = dataFromAppMetrika(start_month, end_month)
    data_ga_device = dataFromGA(analytics, "ga:deviceCategory")
    data_ga_gr = dataFromGA(analytics, "ga:channelGrouping")

    df_app = pd.DataFrame(data_app)
    df_ga_device = pd.DataFrame(data_ga_device)
    df_ga_gr = pd.DataFrame(data_ga_gr)


    metrics = pd.DataFrame({})
    for i, j in appsFlyer.items():
        z = dataFromAppsFlyer(i, j)
        metrics = pd.concat([z, metrics])

    append_df_to_excel(filePath, df_app, sheet_name='MONTH AM', index=False, header=False)
    append_df_to_excel(filePath, df_ga_device, sheet_name='MONTH GA Device', index=False, header=False)
    append_df_to_excel(filePath, df_ga_gr, sheet_name='MONTH GA ChannelGr', index=False, header=False)
    append_df_to_excel(filePath, metrics, sheet_name='MONTH APF metrics', index=False, header=False)

if __name__ == '__main__':
    main()