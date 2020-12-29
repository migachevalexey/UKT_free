import pprint
import requests
import urllib
import datetime
import pandas as pd
from apiclient.discovery import build
from oauth2client.service_account import ServiceAccountCredentials
import httplib2
import urllib3
urllib3.disable_warnings()
from io import StringIO
# print(urllib.parse.unquote('Europe%2fMoscow'))

"""
Тянем данные из GA, из метрики(в разрезе приложения) и из AppsFlyer(в разрезе приложения)
Сохраняем все это в excel с накоплением данных
На еженедельной основе. Работаем через Pandas  
"""

SCOPES = ['https://www.googleapis.com/auth/analytics.readonly']
SERVICE_ACCOUNT_EMAIL = 'ukt-cloud@konic-progress-196909.iam.gserviceaccount.com'
DISCOVERY_URI = 'https://analyticsreporting.googleapis.com/$discovery/rest'
KEY_FILE_LOCATION = r'C:\Python\ukt\pass\uktOwox-d00000000e12.p12'
VIEW_ID = '115850000'  # id "Correct Master Data"

date_start = (datetime.date.today() - datetime.timedelta(days=7)).strftime("%Y-%m-%d")
date_stop = (datetime.date.today() - datetime.timedelta(days=1)).strftime("%Y-%m-%d")
start2weekAgo = (datetime.date.today() - datetime.timedelta(days=14)).strftime("%Y-%m-%d")
stop2weekAgo = (datetime.date.today() - datetime.timedelta(days=8)).strftime("%Y-%m-%d")
print(date_start, date_stop, '----', start2weekAgo, stop2weekAgo)
appsFlyer = {'id582200000': 'iOS', 'ru.ukt.android': 'Android'}
apps = {'com.ak.apps.ukt': 'iOS', 'ru.ukt.android': 'Android'}
filePath = r'\\srv\WEB-ANALYTICS\File_transfer\allDataForReports.xlsx'
weekNumber = (datetime.date.today() - datetime.timedelta(days=3)).strftime("%Y-%V")
weekNumber2 = (datetime.date.today() - datetime.timedelta(days=10)).strftime("%Y-%V")
print(weekNumber, weekNumber2)


def dataFromAppMetrika(d_start, d_stop, weekNum):
    d = []
    for i,j in apps.items():
        params_sess = {'lang': 'ru', 'request_domain': 'ru',
                       'filters': "exists ym:d:device with (appID=='{}')".format(i), 'id': 516000,
                       'date1': d_start, 'date2': d_stop, 'metrics': 'ym:s:sessions', 'dimensions': 'ym:s:date',
                       'sort': '-ym:s:date', 'offset': 1, 'limit': 10, 'accuracy': 1, 'proposedAccuracy': 'true'}
        response_ses = requests.get('https://api.appmetrica.yandex.ru/stat/v1/data', params=params_sess,
                                    headers={'Authorization': 'OAuth AgAAAA'}).json()
        params_users = {'lang': 'ru', 'request_domain': 'ru',
                        'filters': "exists ym:d:device with (appID=='{}')".format(i), 'id': 516000,
                        'date1': d_start, 'date2': d_stop, 'metrics': 'ym:u:activeUsers',
                        'dimensions': 'ym:u:date', 'sort': '-ym:u:date', 'include_undefined': 'true', 'offset': 1,
                        'limit': 10, 'accuracy': 1, 'proposedAccuracy': 'true'}
        response_user = requests.get('https://api.appmetrica.yandex.ru/stat/v1/data', params=params_users,
                                     headers={'Authorization': 'OAuth AgAAAAA'}).json()
        d.append([weekNum, j,  int(*response_ses['totals']),int(*response_user['totals'])])
    return d

dataFromAppMetrika(date_start, date_stop, weekNumber)


def dataFromAppsFlyer(app_id, app_name):
    pd.set_option('display.max_rows', 1000)  # для вывода на  печать полного dataFrame
    pd.set_option('display.max_columns', 20)
    pd.set_option('display.width', 300)
    with open(r'C:\Python\ukt\pass\AppsFlyer.txt') as f:
        tokenApp = f.readline()
    url = f'https://hq.appsflyer.com/export/{app_id}/partners_report/v5'
    param = {'api_token': tokenApp, 'from': date_start, 'to': date_stop, 'timezone': 'Europe/Moscow'}
    response_app = requests.get(url, params=param, verify=False)
    RAWDATA = StringIO(response_app.text)

    df = pd.read_csv(RAWDATA)
    # print(df.head())
    df.columns = [i.replace(' (Unique users)', '') for i in list(df.columns.values)]

    # Вся эту муть на тот случай, если какое-то событие(я) отсутствует. Вычисляем какое именно и добавляем в DataFrame
    primaryColumns={'Media Source (pid)', 'af_add_to_cart', 'af_search', 'af_opened_from_push_notification',
            'af_initiated_checkout', 'af_add_to_wishlist', 'af_rate', 'Impressions', 'Clicks', 'Installs'}
    if not primaryColumns <= set(df.columns.values): # содержит ли недельная выгрузка все события?
        diffColumns = primaryColumns - set(df.columns.values)
        for i in diffColumns:
            df[f'{i}'] = 0

    # df_events = df.filter(['Media Source (pid)', 'af_add_to_cart', 'af_search', 'af_opened_from_push_notification'], axis=1)
    df_first = df.filter(['Media Source (pid)', 'af_add_to_cart', 'af_search', 'af_opened_from_push_notification',
                                            'af_initiated_checkout', 'af_add_to_wishlist','af_rate','Impressions','Clicks','Installs'], axis=1)
    print(df_first.head(5))

    df_first = df_first.groupby(['Media Source (pid)']).sum()
    df_first = df_first.reset_index()
    df_first['system'] = app_name
    df_first['week'] = weekNumber
    df_first['ctr'] = df_first['Installs'] / df_first['Clicks']

    df_events = df_first[['week','system','Media Source (pid)','af_add_to_cart','af_search','af_initiated_checkout','af_opened_from_push_notification','af_add_to_wishlist','af_rate']]

    df_metrics = df_first[['week','system','Media Source (pid)','Impressions','Clicks','Installs','ctr']]

    return df_events, df_metrics

def initialize_analyticsreporting():
    credentials = ServiceAccountCredentials.from_p12_keyfile(
        SERVICE_ACCOUNT_EMAIL, KEY_FILE_LOCATION, scopes=SCOPES)
    http = credentials.authorize(httplib2.Http())
    analytics = build('analytics', 'v4', http=http, discoveryServiceUrl=DISCOVERY_URI)
    return analytics


def dataFromGA(analytics, ga_dim):
    QUERY = {
        'reportRequests': [
            {'viewId': VIEW_ID,
             'dateRanges': [{'startDate': date_start, 'endDate': date_stop}],
             "samplingLevel": 'LARGE',
             "dimensions": [{"name": "ga:isoYearIsoWeek"}, {"name": ga_dim}],
             'metrics': [{"expression": "ga:sessions"}, {"expression": "ga:transactions"},
                         {"expression": "ga:transactionRevenue"}, ],
             'pageSize': 10000}
        ]}
    request = analytics.reports().batchGet(body=QUERY).execute()  # запрос API
    rowCount = request['reports'][0]['data']['rowCount']
    API_data = request['reports'][0]['data']['rows']
    if rowCount > 10000:  # если записей болше 10т, упираемся в лимит и делаем цикл с параметром nextPageToken
        for i in range(rowCount // 10000):
            pageToken = request['reports'][0]["nextPageToken"]
            QUERY['reportRequests'][0].update({"pageToken": pageToken})  # добавляем параметр pageToken в dict запроса
            request = analytics.reports().batchGet(body=QUERY).execute()  # посылаем запрос API
            API_data += request['reports'][0]['data']['rows']

    print(f'Колво записей: {rowCount}')
    API_data = [list(i['dimensions']) + list(i['metrics'][0]['values']) for i in API_data]

    # это надо перенести на уровень обработки  pandas dataFrame
    for i in API_data:
        i[0] = int(i[0])
        i[2] = int(i[2])
        i[3] = int(i[3])
        i[4] = int(float(i[4]))

    return API_data


def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None, truncate_sheet=False, **to_excel_kwargs):
    from openpyxl import load_workbook
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')
    writer = pd.ExcelWriter(filename, engine='openpyxl')
    try:
        writer.book = load_workbook(filename)
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            idx = writer.book.sheetnames.index(sheet_name)
            writer.book.remove(writer.book.worksheets[idx])
            writer.book.create_sheet(sheet_name, idx)
        writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        pass
    if startrow is None:
        startrow = 0
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)
    writer.save()
    writer.close()


def main():
    analytics = initialize_analyticsreporting()
    # вытягиваем данные
    data_app = dataFromAppMetrika(date_start, date_stop, weekNumber) # за прошлую неделю
    data_app2week = dataFromAppMetrika(start2weekAgo, stop2weekAgo, weekNumber2) # за позапрошлую неделю
    data_ga_device = dataFromGA(analytics, "ga:deviceCategory")
    data_ga_gr = dataFromGA(analytics, "ga:channelGrouping")

    # делаем dataFrame
    df_app = pd.DataFrame(data_app)
    df_app2week = pd.DataFrame(data_app2week)
    df_ga_device = pd.DataFrame(data_ga_device)
    df_ga_gr = pd.DataFrame(data_ga_gr)

    events = pd.DataFrame({})
    metrics = pd.DataFrame({})
    for i, j in appsFlyer.items():
        y, z = dataFromAppsFlyer(i, j)
        events = pd.concat([y, events])
        metrics = pd.concat([z, metrics])

    # обрабатываем данные из appMetrica: исторические данные + данные за прошлую и позапрошлую недели
    NewAppDF = pd.concat([df_app2week, df_app]) # данные по последним 2ум неделям
    NewAppDF.columns = ['week', 'app', 'session', 'user']
    xlsx = pd.ExcelFile(filePath)
    historyAppDF = pd.read_excel(xlsx, 'WEEK AM')  # исторические данные
    historyAppDF.drop(historyAppDF.tail(2).index, inplace=True)  # удаляем пердпоследнюю неделю
    FinalAppDF = pd.concat([historyAppDF, NewAppDF])  # полный dataFrame с обновленной предпоследеней неделей

    # добавляем новые данные понедельно
    FinalAppDF.to_csv(filePath[:-22]+'AppDataReports.csv', sep=';', index=False)
    append_df_to_excel(filePath, FinalAppDF, sheet_name='WEEK AM', truncate_sheet=True, index=False, header=True, startrow=0)
    append_df_to_excel(filePath, df_ga_device, sheet_name='WEEK GA Device', index=False, header=False)
    append_df_to_excel(filePath, df_ga_gr, sheet_name='WEEK GA ChannelGr', index=False, header=False)
    append_df_to_excel(filePath, events, sheet_name='WEEK APF events', index=False, header=False)
    append_df_to_excel(filePath, metrics, sheet_name='WEEK APF metrics', index=False, header=False)

main()

