"""Analytics Reporting API V4."""
import pprint

from apiclient.discovery import build
from oauth2client.service_account import ServiceAccountCredentials
import httplib2
import pandas as pd
import matplotlib.pyplot as plt

SCOPES = ['https://www.googleapis.com/auth/analytics.readonly']
SERVICE_ACCOUNT_EMAIL = 'ukt-cloud@konic-progress-196909.iam.gserviceaccount.com'
DISCOVERY_URI = 'https://analyticsreporting.googleapis.com/$discovery/rest'
KEY_FILE_LOCATION = r'C:\Python\ukt\pass\uktOwox-d00000000e12.p12'
VIEW_ID = '115850000'  # id


date_start = '2019-08-18'
date_stop = '2019-09-16'
# date_start = (datetime.date.today() - datetime.timedelta(days=1)).strftime("%Y-%m-%d")
# date_stop=date_start


import numpy as np

pd.set_option('display.max_rows', 1000)  # для печати полного dataFrame
pd.set_option('display.max_columns', 20)
pd.set_option('display.width', 300)

def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None, truncate_sheet=False, **to_excel_kwargs):
    from openpyxl import load_workbook
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')
    writer = pd.ExcelWriter(filename, engine='openpyxl')
    try:
        writer.book = load_workbook(filename)
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            idx = writer.book.sheetnames.index(sheet_name)
            writer.book.remove(writer.book.worksheets[idx])
            writer.book.create_sheet(sheet_name, idx)
        writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        pass
    if startrow is None:
        startrow = 0
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)
    writer.save()

def initialize_analyticsreporting():
    credentials = ServiceAccountCredentials.from_p12_keyfile(
        SERVICE_ACCOUNT_EMAIL, KEY_FILE_LOCATION, scopes=SCOPES)
    http = credentials.authorize(httplib2.Http())
    analytics = build('analytics', 'v4', http=http, discoveryServiceUrl=DISCOVERY_URI)

    return analytics


body = {
        'reportRequests': [
            {'viewId': VIEW_ID,
             'dateRanges': [{'startDate': date_start, 'endDate': date_stop}],
             "samplingLevel": 'LARGE',
             "dimensions": [
                 {"name": "ga:pagePath"},],
             'metrics': [{"expression": "ga:pageviews"}],
             "dimensionFilterClauses": [{
                 "operator": 'AND',
                 "filters": [ # https://developers.google.com/analytics/devguides/reporting/core/v4/rest/v4/reports/batchGet?hl=ru#DimensionFilter
                     {"dimensionName": "ga:pagePath",
                      "expressions": ['/cat/\d+-|/filter|/catalogue_filter'],
                      "operator": "REGEXP"},
                 ]}],
             'pageSize': 10000}
        ]}


def get_report(analytics, QUERY):

    request = analytics.reports().batchGet(body=QUERY).execute()  # запрос API

    rowCount = request['reports'][0]['data'].get('rowCount',0)
    API_data = request['reports'][0]['data']['rows']
    if rowCount > 10000:  # если записей болше 10т, упираемся в лимит и делаем цикл с параметром nextPageToken
        for i in range(rowCount // 10000):
            pageToken = request['reports'][0]["nextPageToken"]
            QUERY['reportRequests'][0].update({"pageToken": pageToken})  # добавляем параметр pageToken в dict запроса
            request = analytics.reports().batchGet(body=QUERY).execute()  # посылаем запрос API
            API_data += request['reports'][0]['data']['rows']
    print(f'Колво записей: {rowCount}')

    API_data = [list(i['dimensions']) + list(i['metrics'][0]['values']) for i in API_data]
    return API_data


def organic_all():
    analytics = initialize_analyticsreporting()

    z_all=get_report(analytics, body)
    # z = get_report(analytics, body)

    df_all = pd.DataFrame.from_records(z_all, columns=['page', 'views']) # 'sourceMedium'
    df_all['views'] = df_all['views'].apply(np.int64)
    df_all=df_all.sort_values(by='views', ascending=False)
    # df['sessions'] = df['sessions'].apply(np.int64)
    # df_final = pd.merge(df, df_all, how='left', on=['date', 'device']).rename(columns={'sessions_x':'sessions', 'sessions_y':'sessions_all'})
    # df_final['ratio'] = df_final['sessions']/df_final['sessions_all']
    df_all.to_csv('../txtcsvFile/Katya2.csv', index=False, sep=';')
    print(df_all.head(10))

    # df_final.to_excel('organic_2018_one filter.xlsx', index=False)

organic_all()

def organic_pages(QUERY, sheetname, period):
    analytics = initialize_analyticsreporting()
    z = get_report(analytics, QUERY)
    df = pd.DataFrame.from_records(z, columns=['date', 'sourceMedium', 'device', 'sess'+period])  #
    df['sess'+period] = df['sess'+period].apply(np.int64)
    append_df_to_excel('organic.xlsx', df, sheet_name=sheetname, index=False)

# for i in [date1, date2]:
#     body['reportRequests'][0]['dateRanges'][0]['startDate']=i[0]
#     body['reportRequests'][0]['dateRanges'][0]['endDate'] = i[1]
#     organic_pages(body, 'action'+i[0], f'{i[0]}/{i[1][5:]}')
